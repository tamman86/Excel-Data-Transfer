import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import os
import math


class ExcelProcessorApp:
    def __init__(self, root_window):
        self.root = root_window
        self.root.title("Datasheet Transfer Tool")
        self.root.geometry("800x700")

        ''' Establish variables '''
        self.source_files = []
        self.base_file = tk.StringVar()
        self.output_folder_name = tk.StringVar()
        self.mapping_rows_data = []                     # Stores info for each mapping in a dictionary

        ''' UI Frames '''
        self.file_selection_frame = ttk.LabelFrame(self.root, text="1. File Selection", padding=10)
        self.file_selection_frame.pack(padx=10, pady=10, fill="x")

        self.mapping_frame_container = ttk.LabelFrame(self.root, text="2. Cell Mappings", padding=10)
        self.mapping_frame_container.pack(padx=10, pady=5, fill="both", expand=True)

        ''' Cell Mappings window '''
        self.canvas = tk.Canvas(self.mapping_frame_container)
        self.scrollbar = ttk.Scrollbar(self.mapping_frame_container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(
                scrollregion=self.canvas.bbox("all")
            )
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        self.action_frame = ttk.LabelFrame(self.root, text="3. Actions", padding=10)
        self.action_frame.pack(padx=10, pady=10, fill="x")

        self.status_frame = ttk.LabelFrame(self.root, text="Status Log", padding=10)
        self.status_frame.pack(padx=10, pady=5, fill="x")

        ''' File Selection window '''
        # Source File(s) Button
        self.select_source_button = ttk.Button(self.file_selection_frame, text="Select Source Excel File(s)",
                                               command=self.select_source_files)
        self.select_source_button.grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        self.source_files_label = ttk.Label(self.file_selection_frame, text="No source files selected.")
        self.source_files_label.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        # Base File Button
        self.select_base_button = ttk.Button(self.file_selection_frame, text="Select Base Excel File",
                                             command=self.select_base_file)
        self.select_base_button.grid(row=1, column=0, padx=5, pady=5, sticky="ew")
        self.base_file_label = ttk.Label(self.file_selection_frame, text="No base file selected.")
        self.base_file_label.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        # Output Folder
        ttk.Label(self.file_selection_frame, text="Output Folder Name:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.output_folder_entry = ttk.Entry(self.file_selection_frame, textvariable=self.output_folder_name, width=40)
        self.output_folder_entry.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

        self.file_selection_frame.columnconfigure(1, weight=1)  # Make label column expandable

        ''' Status Log UI  '''
        self.status_text = tk.Text(self.status_frame, height=5, wrap=tk.WORD, state=tk.DISABLED)
        self.status_text_scrollbar = ttk.Scrollbar(self.status_frame, command=self.status_text.yview)
        self.status_text.config(yscrollcommand=self.status_text_scrollbar.set)
        self.status_text.pack(side="left", fill="both", expand=True)
        self.status_text_scrollbar.pack(side="right", fill="y")

        ''' Initialize Mapping UI '''
        self.add_mapping_row()  # First mapping row

        ''' Mapping Buttons '''
        self.add_row_button = ttk.Button(self.action_frame, text="Add Mapping Row", command=self.add_mapping_row)
        self.add_row_button.pack(side="left", padx=5, pady=5)

        self.transfer_button = ttk.Button(self.action_frame, text="Transfer Values", command=self.transfer_values)
        self.transfer_button.pack(side="left", padx=5, pady=5)

    ''' Status Log window '''
    def log_status(self, message):
        self.status_text.config(state=tk.NORMAL)
        self.status_text.insert(tk.END, message + "\n")
        self.status_text.see(tk.END)  # Scroll to the end
        self.status_text.config(state=tk.DISABLED)
        self.root.update_idletasks()  # Ensure GUI updates

    ''' Source File(s) selection text '''
    def select_source_files(self):
        files = filedialog.askopenfilenames(
            title="Select Source Excel Files",
            filetypes=(("Excel files", "*.xlsx *.xls"), ("All files", "*.*"))
        )
        if files:
            self.source_files = list(files)
            self.source_files_label.config(
                text=f"{len(self.source_files)} file(s) selected: {', '.join([os.path.basename(f) for f in self.source_files])}")
            self.log_status(f"Selected {len(self.source_files)} source file(s).")
        else:
            self.source_files_label.config(text="No source files selected.")
            self.log_status("Source file selection cancelled.")

    ''' Base File(s) selection text '''
    def select_base_file(self):
        file = filedialog.askopenfilename(
            title="Select Base Excel File",
            filetypes=(("Excel files", "*.xlsx *.xls"), ("All files", "*.*"))
        )
        if file:
            self.base_file.set(file)
            self.base_file_label.config(text=f"Base file: {os.path.basename(file)}")
            self.log_status(f"Selected base file: {os.path.basename(file)}")
        else:
            self.base_file_label.config(text="No base file selected.")
            self.log_status("Base file selection cancelled.")

    ''' Adding Mapping Rows '''
    def add_mapping_row(self):
        row_frame = ttk.Frame(self.scrollable_frame, padding=5)
        row_frame.pack(fill="x", pady=2)

        row_number = len(self.mapping_rows_data) + 1
        ttk.Label(row_frame, text=f"Mapping {row_number}:").grid(row=0, column=0, columnspan=6, sticky="w", pady=(0, 5))

        # "From" row
        ttk.Label(row_frame, text="From:").grid(row=1, column=0, padx=2, sticky="w")
        ttk.Label(row_frame, text="Row:").grid(row=1, column=1, padx=2)
        from_row_var = tk.StringVar()
        from_row_entry = ttk.Entry(row_frame, textvariable=from_row_var, width=5)
        from_row_entry.grid(row=1, column=2, padx=2)

        ttk.Label(row_frame, text="Col:").grid(row=1, column=3, padx=2)
        from_col_var = tk.StringVar()
        from_col_entry = ttk.Entry(row_frame, textvariable=from_col_var, width=5)
        from_col_entry.grid(row=1, column=4, padx=2)

        # "To" row
        ttk.Label(row_frame, text="To:").grid(row=2, column=0, padx=2, sticky="w")
        ttk.Label(row_frame, text="Row:").grid(row=2, column=1, padx=2)
        to_row_var = tk.StringVar()
        to_row_entry = ttk.Entry(row_frame, textvariable=to_row_var, width=5)
        to_row_entry.grid(row=2, column=2, padx=2)

        ttk.Label(row_frame, text="Col:").grid(row=2, column=3, padx=2)
        to_col_var = tk.StringVar()
        to_col_entry = ttk.Entry(row_frame, textvariable=to_col_var, width=5)
        to_col_entry.grid(row=2, column=4, padx=2)

        # Conversion Option
        convert_var = tk.BooleanVar()
        formula_var = tk.StringVar()
        formula_entry = ttk.Entry(row_frame, textvariable=formula_var, width=20, state=tk.DISABLED)

        def toggle_formula_entry():
            if convert_var.get():
                formula_entry.config(state=tk.NORMAL)
                test_button.config(state=tk.NORMAL)
            else:
                formula_entry.config(state=tk.DISABLED)
                test_button.config(state=tk.DISABLED)
                demo_label.config(text="")

        convert_check = ttk.Checkbutton(row_frame, text="Convert", variable=convert_var, command=toggle_formula_entry)
        convert_check.grid(row=3, column=0, columnspan=2, padx=2, sticky="w")

        ttk.Label(row_frame, text="Formula (use 'X'):").grid(row=3, column=2, columnspan=2, padx=2, sticky="e")
        formula_entry.grid(row=3, column=4, columnspan=2, padx=2, sticky="ew")

        # Test conversion formula
        demo_text_var = tk.StringVar(value="X=1 -> ?")
        demo_label = ttk.Label(row_frame, textvariable=demo_text_var, width=20)
        demo_label.grid(row=4, column=2, columnspan=2, padx=2, pady=(2, 0), sticky="w")

        # Create convert test button
        test_button = ttk.Button(
            row_frame,
            text="Test Formula",
            command=lambda fv=formula_var, dl=demo_label: (self.test_formula_conversion(fv, dl))
        )
        test_button.grid(row=4, column=0, columnspan=2, padx=2, pady=(2, 0), sticky="w")

        # Test button initially disabled
        if not convert_var.get():
            test_button.config(state=tk.DISABLED)
            demo_label.config(text="")  # Clear demo text if convert is off

        row_frame.columnconfigure(5, weight=1)  # Make formula entry expandable

        # Store variables in dictionary
        self.mapping_rows_data.append({
            "from_row": from_row_var, "from_col": from_col_var,
            "to_row": to_row_var, "to_col": to_col_var,
            "convert": convert_var, "formula": formula_var
        })

        self.log_status(f"Added mapping row {row_number}.")
        self.scrollable_frame.update_idletasks()
        self.canvas.config(scrollregion=self.canvas.bbox("all"))

    def test_formula_conversion(self, formula_var, demo_label_widget):
        formula_str = formula_var.get()
        original_formula_for_test_log = formula_str
        formula_str = formula_str.replace('x', 'X')  # Convert 'x' to 'X' for both cases
        if not formula_str:
            demo_label_widget.config(text="X=1 -> (enter formula)")
            return
        try:
            eval_globals = {"__builtins__": {}, "math": math}
            eval_locals = {"X": 1}  # Test with X=1

            result = eval(formula_str, eval_globals, eval_locals)
            print(result)
            # Format result nicely, especially floats
            if isinstance(result, float):
                demo_text = f"X=1 -> {result:.4f}"
            else:
                demo_text = f"X=1 -> {result}"
            demo_label_widget.config(text=demo_text, foreground="green")
            self.log_status(f"Formula Test (X=1): User entered '{original_formula_for_test_log}', Evaluated as '{formula_str}' -> Result: {result}")
        except Exception as e:
            demo_label_widget.config(text="Invalid equation", foreground="red")
            self.log_status(f"Formula Test (X=1): User entered '{original_formula_for_test_log}', Evaluated as '{formula_str}' -> Invalid equation. Error: {e}")

    # Move values over
    def transfer_values(self):
        if not self.source_files:
            messagebox.showerror("Error", "Please select at least one source Excel file.")
            return
        if not self.base_file.get():
            messagebox.showerror("Error", "Please select a base Excel file.")
            return
        if not self.mapping_rows_data:  # Check for dictionary or dictionaries
            # Need all From and To info
            if not any(self.mapping_rows_data[0][key].get() for key in ["from_row", "from_col", "to_row", "to_col"]):
                messagebox.showerror("Error", "Please add and fill at least one mapping row.")
                return

        output_folder_str = self.output_folder_name.get().strip()
        if not output_folder_str:
            messagebox.showerror("Error", "Please specify an output folder name.")
            self.log_status("Error: Output folder name not specified.")
            return

        # Create output folder in same directory as base file
        base_file_full_path = self.base_file.get()
        if not base_file_full_path or not os.path.isfile(base_file_full_path):
            messagebox.showerror("Error",
                                 "Base file path is invalid or not selected. Cannot determine output folder location.")
            self.log_status("Error: Base file path invalid for output folder creation.")
            return

        # Put folder in same folder as Base File
        base_file_dir = os.path.dirname(base_file_full_path)
        output_folder_path = os.path.join(base_file_dir, output_folder_str)
        output_folder_path = os.path.abspath(output_folder_path)

        try:
            if not os.path.exists(output_folder_path):
                os.makedirs(output_folder_path)
                self.log_status(f"Created output folder: {output_folder_path}")
            else:
                if not os.path.isdir(output_folder_path):
                    messagebox.showerror("Error", f"Output path '{output_folder_path}' exists but is not a folder.")
                    self.log_status(f"Error: Output path '{output_folder_path}' is not a folder.")
                    return
                self.log_status(f"Using existing output folder: {output_folder_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not create or access output folder '{output_folder_path}': {e}")
            self.log_status(f"Error creating/accessing output folder '{output_folder_path}': {e}")
            return

        self.log_status("Starting value transfer process...")
        processed_files_count = 0

        # Get From - To info out of dictionary
        mappings = []
        for i, row_data in enumerate(self.mapping_rows_data):
            try:
                # Check for From and To
                if not (row_data["from_row"].get() or row_data["from_col"].get() or \
                        row_data["to_row"].get() or row_data["to_col"].get()):
                    if len(self.mapping_rows_data) == 1:  # If it's the only row and it's empty
                        messagebox.showerror("Input Error",
                                             "The mapping row is empty. Please fill in the row and column numbers.")
                        self.log_status("Error: Mapping row is empty.")
                        return
                    continue  # Skip this empty row if there are other rows

                from_r_str = row_data["from_row"].get()
                from_c_str = letter_convert[row_data["from_col"].get().lower()]
                to_r_str = row_data["to_row"].get()
                to_c_str = letter_convert[row_data["to_col"].get().lower()]

                if not (from_r_str and from_c_str and to_r_str and to_c_str):
                    messagebox.showerror("Input Error", f"Missing Row/Column in mapping row {i + 1}.")
                    self.log_status(f"Error in mapping row {i + 1}: Missing Row/Column.")
                    return

                from_r = int(from_r_str)
                from_c = int(from_c_str)
                to_r = int(to_r_str)
                to_c = int(to_c_str)

                if not (from_r > 0 and from_c > 0 and to_r > 0 and to_c > 0):
                    raise ValueError("Invalid Row/Column specified.")

                mappings.append({
                    "from_row": from_r, "from_col": from_c,
                    "to_row": to_r, "to_col": to_c,
                    "convert": row_data["convert"].get(),
                    "formula": row_data["formula"].get()
                })
            except ValueError as e:
                messagebox.showerror("Input Error",
                                     f"Invalid input in mapping row {i + 1}: {e}\nPlease enter valid positive integers for rows and columns.")
                self.log_status(f"Error in mapping row {i + 1}: {e}")
                return

        if not mappings:  # If after validation, no valid mappings were collected
            messagebox.showerror("Error",
                                 "No valid mappings provided. Please fill in at least one mapping row correctly.")
            self.log_status("Error: No valid mappings collected.")
            return

        self.log_status(f"Collected {len(mappings)} mapping configurations.")

        # Iterate Source Files
        for source_file_path in self.source_files:
            try:
                self.log_status(f"\nProcessing source file: {os.path.basename(source_file_path)}")

                # Load source workbook
                source_wb = openpyxl.load_workbook(source_file_path, data_only=True)
                source_sheet = source_wb.active

                # Load a fresh copy of the base workbook for each source file
                base_wb_path = self.base_file.get()
                base_wb_copy = openpyxl.load_workbook(base_wb_path)
                base_sheet_copy = base_wb_copy.active

                # Apply mappings
                for i, mapping in enumerate(mappings):
                    self.log_status(
                        f"  Applying mapping {i + 1}: From ({mapping['from_row']},{mapping['from_col']}) To ({mapping['to_row']},{mapping['to_col']})")

                    try:
                        source_value = source_sheet.cell(row=mapping["from_row"], column=mapping["from_col"]).value
                        self.log_status(
                            f"    Read value '{source_value}' from source cell ({mapping['from_row']},{mapping['from_col']}).")
                    except Exception as e:
                        self.log_status(
                            f"    ERROR reading from source cell ({mapping['from_row']},{mapping['from_col']}): {e}")
                        continue

                    value_to_paste = source_value

                    if mapping["convert"] and mapping["formula"]:
                        formula_str = mapping["formula"]
                        original_formula_for_log = formula_str  # Keep original for logging
                        formula_str = formula_str.replace('x', 'X')  # Convert 'x' to 'X' for both cases
                        try:
                            eval_globals = {"__builtins__": {}, "math": math}
                            eval_locals = {"X": source_value}
                            converted_value = eval(formula_str, eval_globals, eval_locals)
                            value_to_paste = converted_value
                            self.log_status(f"    Applied formula (User: '{original_formula_for_log}', Evaluated: '{formula_str}'). Original: {source_value}, Converted: {converted_value}")
                        except Exception as e:
                            self.log_status(f"    ERROR applying formula (User: '{original_formula_for_log}', Evaluated: '{formula_str}') to value '{source_value}': {e}. Using original value.")

                    try:
                        base_sheet_copy.cell(row=mapping["to_row"], column=mapping["to_col"]).value = value_to_paste
                        self.log_status(
                            f"    Wrote value '{value_to_paste}' to target cell ({mapping['to_row']},{mapping['to_col']}).")
                    except Exception as e:
                        self.log_status(
                            f"    ERROR writing to target cell ({mapping['to_row']},{mapping['to_col']}): {e}")

                # Define output path within the specified output folder
                output_filename = os.path.join(output_folder_path, os.path.basename(source_file_path))

                try:
                    base_wb_copy.save(output_filename)
                    self.log_status(f"  Successfully processed. Output saved as: {output_filename}")
                    processed_files_count += 1
                except PermissionError:
                    self.log_status(
                        f"  PERMISSION ERROR saving processed file {output_filename}. Check folder permissions.")
                except Exception as e_save:
                    self.log_status(f"  ERROR saving processed file {output_filename}: {e_save}")


            except FileNotFoundError:
                self.log_status(f"ERROR: Source file not found: {source_file_path}")
            except Exception as e:
                self.log_status(
                    f"An unexpected error occurred while processing {os.path.basename(source_file_path)}: {e}")

        self.log_status(f"\n--- Transfer Complete ---")
        self.log_status(
            f"Successfully processed {processed_files_count} out of {len(self.source_files)} source file(s).")
        if processed_files_count > 0:
            messagebox.showinfo("Success",
                                f"Successfully processed {processed_files_count} file(s). Check the log for details and output locations.")
        elif self.source_files:  # Only show error if there were files to process
            messagebox.showerror("Processing Issue",
                                 "No files were processed successfully, or there were issues saving. Check the log for errors.")


if __name__ == '__main__':

    ''' Create dictionary for column conversion '''
    letter_convert = {}
    count = 1

    for i in range(ord('a'), ord('z') + 1):
        letter_convert[chr(i)] = count
        count += 1

    for i in range(ord('a'), ord('z') + 1):
        for j in range(ord('a'), ord('z') + 1):
            two_letter = chr(i) + chr(j)
            letter_convert[two_letter] = count
            count += 1

    main_root = tk.Tk()
    app = ExcelProcessorApp(main_root)
    main_root.mainloop()
